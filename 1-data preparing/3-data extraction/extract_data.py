import pdfplumber
import re
import os
from openpyxl import Workbook, load_workbook
import time
from datetime import datetime

# 📁 Chemin vers le PDF et le fichier Excel
pdf_path = "C:/Users/lenovo/Downloads/stageapplication2/pdfs/PDF_fusionne_total.pdf"
excel_path = "C:/Users/lenovo/Downloads/stageapplication2/data extraction/annonces_marches_publics.xlsx"

# 📊 Variables de monitoring
total_blocs_traites = 0
total_extractions_reussies = 0
total_echecs = 0
debut_extraction = time.time()

# 📄 Initialisation du fichier Excel s'il n'existe pas
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Marchés Publics"
    ws.append([
        "Procédure", "Catégorie", "Publié le", "Référence",
        "Objet", "Acheteur public", "Lieu d'exécution"
    ])
    wb.save(excel_path)

# 🔄 Ouvrir le fichier Excel existant
wb = load_workbook(excel_path)
ws = wb.active

def afficher_progression(page_actuelle, total_pages, blocs_page, bloc_actuel):
    """Affiche la progression détaillée"""
    temps_ecoule = time.time() - debut_extraction
    
    print(f"\n{'='*60}")
    print(f"⏰ TEMPS ÉCOULÉ: {temps_ecoule:.1f}s")
    print(f"📄 PAGE: {page_actuelle}/{total_pages} ({(page_actuelle/total_pages)*100:.1f}%)")
    print(f"🔍 BLOC: {bloc_actuel}/{blocs_page}")
    print(f"📊 TOTAL TRAITÉS: {total_blocs_traites}")
    print(f"✅ SUCCÈS: {total_extractions_reussies}")
    print(f"❌ ÉCHECS: {total_echecs}")
    if total_blocs_traites > 0:
        print(f"🎯 TAUX DE RÉUSSITE: {(total_extractions_reussies/total_blocs_traites)*100:.1f}%")
    print(f"{'='*60}")

def afficher_donnees_extraites(donnees, bloc_num):
    """Affiche les données extraites pour validation"""
    print(f"\n📋 BLOC {bloc_num} - DONNÉES EXTRAITES:")
    print(f"  🏢 Procédure: [{donnees['procedure']}]")
    print(f"  📂 Catégorie: [{donnees['categorie']}]")
    print(f"  📅 Publié le: [{donnees['publie_le']}]")
    print(f"  🔢 Référence: [{donnees['reference']}]")
    print(f"  📝 Objet: [{donnees['objet'][:80]}{'...' if len(donnees['objet']) > 80 else ''}]")
    print(f"  🏛️ Acheteur: [{donnees['acheteur'][:50]}{'...' if len(donnees['acheteur']) > 50 else ''}]")
    print(f"  📍 Lieu: [{donnees['lieu']}]")

def nettoyer_texte(texte):
    """Nettoie un texte en supprimant les éléments parasites"""
    if not texte:
        return ""
    
    # Supprimer les mots-clés de structure
    mots_cles_a_supprimer = ['Publié le', 'Référence', 'Objet', 'Acheteur public', 'Lieu d\'exécution', 'Type d\'annonce', 'Détail', 'Actions']
    texte_nettoye = texte
    
    for mot in mots_cles_a_supprimer:
        texte_nettoye = re.sub(rf'\b{re.escape(mot)}\b\s*:?\s*', '', texte_nettoye, flags=re.IGNORECASE)
    
    # Nettoyer les espaces multiples et retours à la ligne
    texte_nettoye = re.sub(r'\s+', ' ', texte_nettoye)
    
    return texte_nettoye.strip()

def extraire_lieu_robuste(bloc):
    """Extrait le lieu d'exécution avec plusieurs méthodes"""
    
    # Liste des villes marocaines connues
    villes_maroc = [
        'CASABLANCA', 'RABAT', 'FES', 'MEKNES', 'MARRAKECH', 'AGADIR', 'TANGER', 'OUJDA', 
        'KENITRA', 'TETOUAN', 'SALE', 'TEMARA', 'MOHAMMEDIA', 'KHOURIBGA', 'JADIDA', 
        'BENI MELLAL', 'NADOR', 'BERKANE', 'TAOURIRT', 'OUEZZANE', 'CHEFCHAOUEN', 
        'LARACHE', 'KSAR KEBIR', 'SIDI KACEM', 'SIDI SLIMANE', 'OUARZAZATE', 'ZAGORA', 
        'ERRACHIDIA', 'FQUIH BEN SALAH', 'KHENIFRA', 'IFRANE', 'AZROU', 'MIDELT', 
        'BOULEMANE', 'TAOUNATE', 'TAZA', 'GUERCIF', 'JERADA', 'FIGUIG', 'ESSAOUIRA', 
        'SAFI', 'YOUSSOUFIA', 'KALAA SRAGHNA', 'CHICHAOUA', 'TAROUDANT', 'TIZNIT', 
        'GUELMIM', 'TAN TAN', 'LAAYOUNE', 'DAKHLA', 'SMARA', 'BOUJDOUR', 'ANGAD',
        'CARDOE', 'ABHGZR', 'NOUACEUR', 'ASSILAH', 'SONARGES'
    ]
    
    lignes = bloc.strip().split('\n')
    
    # Méthode 1: Chercher dans la structure tabulaire (dernière colonne)
    for ligne in lignes:
        colonnes = re.split(r'\s{2,}', ligne.strip())
        if len(colonnes) >= 2:
            derniere_colonne = colonnes[-1].strip()
            if derniere_colonne.upper() in villes_maroc:
                return derniere_colonne.upper()
    
    # Méthode 2: Chercher les villes connues dans tout le bloc
    for ville in villes_maroc:
        if re.search(rf'\b{re.escape(ville)}\b', bloc, re.IGNORECASE):
            return ville
    
    # Méthode 3: Chercher des patterns de lieu en fin de ligne
    for ligne in reversed(lignes[-5:]):  # Chercher dans les 5 dernières lignes
        ligne = ligne.strip()
        if (3 <= len(ligne) <= 30 and 
            ligne.isupper() and 
            not re.search(r'\d{2}/\d{2}/\d{4}', ligne) and
            not re.search(r'(AOO|AOS|CONCA|APPEL|AVIS|Services|Travaux|Fournitures)', ligne)):
            return ligne
    
    return ""

def extraire_donnees_bloc_robuste(bloc):
    """Extrait les données avec une approche plus robuste"""
    
    # Initialiser les variables
    donnees = {
        'procedure': '',
        'categorie': '',
        'publie_le': '',
        'reference': '',
        'objet': '',
        'acheteur': '',
        'lieu': '',
    }
    
    # Nettoyer le bloc
    bloc_clean = bloc.strip()
    
    # 1. PROCÉDURE - Chercher au début du bloc
    procedure_match = re.search(r'^(AOO|AOS|CONCA|APPEL|AVIS)', bloc_clean, re.MULTILINE)
    if procedure_match:
        donnees['procedure'] = procedure_match.group(1)
    else:
        # Si pas trouvé au début, chercher dans tout le bloc
        procedure_match = re.search(r'\b(AOO|AOS|CONCA|APPEL|AVIS)\b', bloc_clean)
        if procedure_match:
            donnees['procedure'] = procedure_match.group(1)
    
    # 2. CATÉGORIE
    categorie_match = re.search(r'\b(Services|Travaux|Fournitures)\b', bloc_clean)
    if categorie_match:
        donnees['categorie'] = categorie_match.group(1)
    
    # 3. DATE DE PUBLICATION
    date_match = re.search(r'\b\d{2}/\d{2}/\d{4}\b', bloc_clean)
    if date_match:
        donnees['publie_le'] = date_match.group(0)
    
    # 4. RÉFÉRENCE
    reference_patterns = [
        r'\b\d{1,4}[/\-_][A-Z]{2,}[/\-_][A-Z]{2,}[/\-_]\d{4}\b',  # Format: 028/DPO/INV/2025
        r'\b\d{1,4}[/\-_]\d{4}[/\-_][A-Z]{2,}\b',  # Format: 39/2025/SMG
        r'\b\d{1,4}[_][A-Z]{2,}[_][A-Z]{2,}[_]\d{4}\b',  # Format: 030_DPO_INV_2025
        r'\b\d{1,4}[/\-_]\d{4}\b',  # Format simple: 123/2025
    ]
    
    for pattern in reference_patterns:
        ref_match = re.search(pattern, bloc_clean)
        if ref_match:
            donnees['reference'] = ref_match.group(0)
            break
    
    # 5. OBJET - Chercher après "Objet :"
    objet_patterns = [
        r'Objet\s*:\s*(.+?)(?=\nAcheteur|$)',
        r'Objet\s*:\s*(.+?)(?=\n[A-Z][a-z]+|$)',
        r'Objet\s*:\s*(.+?)(?=\n\s*\n|$)',
    ]
    
    for pattern in objet_patterns:
        objet_match = re.search(pattern, bloc_clean, re.DOTALL)
        if objet_match:
            donnees['objet'] = nettoyer_texte(objet_match.group(1))
            break
    # ✅ Supprimer toute date dans l'objet, s'il existe
    donnees['objet'] = re.sub(r'\d{2}/\d{2}/\d{4}', '', donnees['objet']).strip()

    # Si pas d'objet trouvé avec "Objet :", chercher le texte principal
    if not donnees['objet']:
        # Chercher le texte le plus long qui ne contient pas de mots-clés
        lignes = bloc_clean.split('\n')
        for ligne in lignes:
            ligne = ligne.strip()
            if (len(ligne) > 20 and 
                not re.search(r'(AOO|AOS|CONCA|APPEL|AVIS|Services|Travaux|Fournitures|\d{2}/\d{2}/\d{4})', ligne) and
                not re.search(r'(Publié le|Référence|Acheteur public|Lieu d\'exécution)', ligne)):
                donnees['objet'] = nettoyer_texte(ligne)
                break
    # ✅ Supprimer à nouveau la date si présente dans la seconde méthode
    donnees['objet'] = re.sub(r'\d{2}/\d{2}/\d{4}', '', donnees['objet']).strip()
   
    # 6. ACHETEUR PUBLIC
    acheteur_patterns = [
        r'Acheteur public\s*:\s*(.+?)(?=\n[A-Z]{3,}|$)',
        r'Acheteur\s*:\s*(.+?)(?=\n[A-Z]{3,}|$)',
        r'Organisme\s*:\s*(.+?)(?=\n[A-Z]{3,}|$)',
    ]
    
    for pattern in acheteur_patterns:
        acheteur_match = re.search(pattern, bloc_clean, re.DOTALL)
        if acheteur_match:
            donnees['acheteur'] = nettoyer_texte(acheteur_match.group(1))
            break
    
    # 7. LIEU D'EXÉCUTION
    donnees['lieu'] = extraire_lieu_robuste(bloc_clean)
    # 🚫 Supprimer l'objet s'il contient une date
    

    return donnees

# 🚀 DÉBUT DE L'EXTRACTION
print(f"🚀 DÉBUT DE L'EXTRACTION - {datetime.now().strftime('%H:%M:%S')}")
print(f"📁 Fichier PDF: {pdf_path}")
print(f"📊 Fichier Excel: {excel_path}")

# 📄 Lecture du PDF
with pdfplumber.open(pdf_path) as pdf:
    total_pages = len(pdf.pages)
    print(f"📄 NOMBRE TOTAL DE PAGES: {total_pages}")
    
    for page_num, page in enumerate(pdf.pages, 1):
        text = page.extract_text()
        if not text:
            print(f"⚠️ Page {page_num} vide - ignorée")
            continue

        # Diviser en blocs par procédure ET par saut de ligne important
        blocs = re.split(r"\n(?=AOO|AOS|CONCA|APPEL|AVIS)", text)
        
        # Filtrer les blocs trop courts mais être moins strict
        blocs_valides = [b for b in blocs if b.strip() and len(b.strip()) >= 20]
        
        print(f"\n📄 PAGE {page_num}/{total_pages} - {len(blocs_valides)} blocs détectés")

        for i, bloc in enumerate(blocs_valides, 1):
            total_blocs_traites += 1
            
            # Afficher la progression toutes les 10 extractions
            if total_blocs_traites % 10 == 0:
                afficher_progression(page_num, total_pages, len(blocs_valides), i)
            
            # Extraire les données du bloc
            donnees = extraire_donnees_bloc_robuste(bloc)
            
            # Critères de validation plus souples
            extraction_valide = (
                # Au moins une procédure OU un objet long
                (donnees['procedure'] or len(donnees['objet']) > 10) and
                # Au moins l'objet OU l'acheteur
                (donnees['objet'] or donnees['acheteur'])
            )
            
            if extraction_valide:
                total_extractions_reussies += 1
                
                # Afficher les données extraites
                afficher_donnees_extraites(donnees, total_blocs_traites)
                
                # Ajout à Excel
                ws.append([
                    donnees['procedure'],
                    donnees['categorie'],
                    donnees['publie_le'],
                    donnees['reference'],
                    donnees['objet'],
                    donnees['acheteur'],
                    donnees['lieu'],
                ])
                
                print(f"  ✅ AJOUTÉ À EXCEL")
            else:
                total_echecs += 1
                print(f"  ❌ ÉCHEC - Données insuffisantes")
                print(f"     Procédure: [{donnees['procedure']}]")
                print(f"     Objet: [{donnees['objet'][:50]}...]")
                print(f"     Acheteur: [{donnees['acheteur'][:30]}...]")
                print(f"     Lieu: [{donnees['lieu']}]")
                
                # DEBUG: Afficher un échantillon du bloc pour comprendre
                print(f"     ÉCHANTILLON BLOC: [{bloc[:100]}...]")
            
            # Petit délai pour pouvoir suivre l'extraction
            time.sleep(0.05)  # Réduit pour accélérer

        # Sauvegarde après chaque page
        wb.save(excel_path)
        print(f"💾 PAGE {page_num} SAUVEGARDÉE")

# 📊 RÉSUMÉ FINAL
temps_total = time.time() - debut_extraction
print(f"\n{'='*70}")
print(f"🎉 EXTRACTION TERMINÉE - {datetime.now().strftime('%H:%M:%S')}")
print(f"⏰ TEMPS TOTAL: {temps_total:.1f} secondes")
print(f"📄 PAGES TRAITÉES: {total_pages}")
print(f"📊 BLOCS TRAITÉS: {total_blocs_traites}")
print(f"✅ EXTRACTIONS RÉUSSIES: {total_extractions_reussies}")
print(f"❌ ÉCHECS: {total_echecs}")
print(f"🎯 TAUX DE RÉUSSITE: {(total_extractions_reussies/total_blocs_traites)*100:.1f}%")
print(f"💾 FICHIER EXCEL: {excel_path}")
print(f"{'='*70}")
